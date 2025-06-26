from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import logging

logger = logging.getLogger("my_app")


def clean_xlsx(file_path: str) -> None:

    logger.info(f"Cleaning Excel file: {file_path}")

    try:
        wb = load_workbook(file_path)

        fonte_preta = Font(color="000000")

        fundo_branco = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid"
        )

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = fonte_preta
                    cell.fill = fundo_branco

        wb.save(file_path)
        logger.info(f"Excel file cleaned successfully: {file_path}")

    except Exception:
        logger.exception(f"Error cleaning Excel file: {file_path}")
